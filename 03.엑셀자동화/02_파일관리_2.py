import  openpyxl

save_path = './temp/자동화excel.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path)

# 활성화된 시트 선태
ws = wb.active

# 데이터 추가 (1)
ws['A1'] = '날짜'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'
ws['E1'] = '합계'

#데이터 추가 (2)
ws.cell(row=2, column=1, value='2022.05.16')
ws.cell(row=2, column=2, value='갤럭시 s22')
ws.cell(row=2, column=3, value=999000)
ws.cell(row=2, column=4, value=2)
ws.cell(row=2, column=5, value='=C2*D2')

#데이타 추가 (3)
ws.append(['2022.05.17', '갤럭시 S21', 550000, 3, '=C3*D3'])

# 데이터 삭제
del ws['A3']

#엑셀저장
wb.save(save_path)