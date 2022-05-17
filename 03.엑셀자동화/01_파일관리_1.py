import openpyxl

# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 현재 활성화된 시트 선택
ws = wb.active

# Sheet 이름 변경
ws.title = '자동화Sheet1'

# 새로운 시트 생성
ws1 = wb.create_sheet('2021.05')

print(wb.sheetnames)

# Sheet 삭제
del wb['자동화Sheet1']

# 엑셀 저장
wb.save('./temp/자동화excel.xlsx')