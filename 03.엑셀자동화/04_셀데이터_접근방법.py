import openpyxl

save_path = './temp/스마트스토어.xlsx'

# 기본 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path, data_only=True) # data_only=True =  수식이 걸린 셀의 값을 가져올 때

# data 시트 선택
ws = wb['data']

# 01. 모든 셀 데이터 가져오기
# -> 행과 열 개수를 아는 경우
# for x in range(1, 11 + 1):
#     for y in range(1, 5 + 1):
#         print(ws.cell(row=x, column=y).value, end= " ") # end= " " 한칸 띄우기
#     print()

# -> 행과 열 개수를 모르는 경우
# for x in range(1, ws.max_row + 1): # max_row = 최대 행 수
#     for y in range(1, ws.max_column + 1): # max_column = 최대 열 수
#         print(ws.cell(row=x, column=y).value, end= " ") # end= " " 한칸 띄우기
#     print()

# 모든 행 가져오기
# for row in ws.iter_rows():
#     print(row)

# 2번째 행 부터 가져오기
# for row in ws.iter_rows(min_row=2):
#     print(row)

# 2번째 행 부터 5번째 행 가져오기
# for row in ws.iter_rows(min_row=2, max_row=5):
#     print(row)

# 2-4행, 2-4열
for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4):
    for cell in row:
        print(cell.value, end=" ")
    print()